# -*- coding:utf-8 -*-
# 把某一模块在固定时间内发送的全部数据提取出来

import re
import openpyxl
import csv

global imei, date1

date1 = ''
imei = '866545052217954'  # 车载型


def split_file(file_name):
    GPS = []  # 存放gps+lbs
    gps_time = []  # 存放(系统时间,设备时间,设备启动/熄火状态）
    GPS_1 = []  # 存放gps
    OB = []  # 存放OB
    num1 = 0  # GPS计数
    num3 = 0  # LBS计数
    num = 0  # 总计数

    # with open(file_name, 'r') as f:
    # reader = csv.reader(f)
    # 数据的读取
    file = open(file_name, encoding='utf-8')
    line_content = file.readlines()
    process_content_list = list()
    process_content_list.append(['@timestamp', 'message'])
    for item in line_content:
        temp = item.strip()
        line_list = list()
        if len(temp) != 0:
            line_list.append(temp[0:28])
            line_list.append(temp[29:])
            process_content_list.append(line_list)
    for row in process_content_list:
        # if re.search(date1 + r'.*>0A,[1-3],' + imei, row[1]) != None :
        if re.search(r'.*>0A,[1-3],' + imei, row[1]) != None:
            split_line = row[1].split(' ')
            a = split_line[0] + ' ' + (split_line[1])[:8]
            add_time1 = split_line[0] + ' ' + (split_line[1])[:8]  # 平台的时间
            for i in split_line[6:]:
                a = a + ' ' + i  # a 为 设备上传时间+originals: 后面的内容
            add_time2 = (split_line[7])[:8]  # 设备的时间
            car_state = (split_line[7])[11:14]  # 设备的电池状态+启动/熄火
            num = num + 1
            gps_time.append((add_time1, add_time2, car_state, a))

            # 判断该行是否是GPS数据，是的话单独保存GPS数据并计数
            if re.search(r'>0A,1,' + imei, a) != None:
                num1 = num1 + 1
                a = a + '/n'
                GPS_1.append(a)
                GPS.append(a)
            elif re.search(r'>0A,3,' + imei, a) != None:
                num3 = num3 + 1
                GPS.append(a)
        # 提取注册包0B
        elif re.search(r'>0B,' + imei, row[1]) != None:
            split_line = row[1].split(' ')
            a = split_line[0] + ' ' + (split_line[1])[:8]
            for i in split_line[6:]:
                a = a + ' ' + i
            OB.append(a)

    print('总条数：%d' % num)
    print('GPS条数：%d' % num1)
    print('LBS条数：%d' % num3)

    data = imei[-4:]
    # 存GPS+LBS的文件
    file_name_GPS = data + '.txt'
    GPS_file = open(r'G:\\工作文件\\线装式在途定位器\\生成数据\\' + file_name_GPS, 'w')
    # 存GPS的文件
    file_name_GPS1 = data + '_gps' + '.txt'
    GPS1_file = open(r'G:\\工作文件\\线装式在途定位器\\生成数据\\' + file_name_GPS1, 'w')
    # 存0B的文件
    file_name_OB = data + '_OB' + '.txt'
    OB_file = open(r'G:\\工作文件\\线装式在途定位器\\生成数据\\' + file_name_OB, 'w')

    # 把GPS数据的列表内容写入文件
    GPS_file.writelines('\n'.join(GPS))
    GPS_file.close()
    GPS1_file.writelines('\n'.join(GPS_1))
    GPS1_file.close()
    OB_file.writelines('\n'.join(OB))
    OB_file.close()
    # f.close()

    # 把时间写入表格
    book = openpyxl.Workbook()
    sh = book.active
    row = 1
    gps_time = gps_time[::-1]
    for t1, t2, yd, a in gps_time:
        sh.cell(row, 1).value = t1
        sh.cell(row, 3).value = t2
        sh.cell(row, 5).value = yd
        sh.cell(row, 7).value = a
        row += 1
    # row = 1
    # for g in g4:
    #     sh.cell(row, 5).value = g
    #     row += 1
    excel_name = 'G:\\工作文件\\线装式在途定位器\\生成数据\\' + data + '.xlsx'
    book.save(excel_name)


split_file('G:\\工作文件\\线装式在途定位器\\数据\\原始数据.txt')
#    split_file('C:\Users\260386\Downloads\新保存的搜索.csv')
